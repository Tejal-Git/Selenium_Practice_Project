import openpyxl
from selenium import webdriver
import time
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()
driver.maximize_window()

file="D:\Tejal\software projects\Selenium_Projects\Git\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook[("Sheet1")]

rows= sheet.max_row
cols=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end="     ")
