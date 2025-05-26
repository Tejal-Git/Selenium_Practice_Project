import openpyxl

file= "D:\Tejal\software projects\Selenium_Projects\Git\Book2.xlsx"

# workbook=openpyxl.load_workbook(file)
# sheet=workbook.active
#
# for r in range (1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="Welcome"
#
#
# workbook.save(file)

workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value=1
sheet.cell(1,2).value="Akash"

sheet.cell(2,1).value=2
sheet.cell(2,2).value="Tejal"
workbook.save(file)
